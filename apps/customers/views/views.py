from rest_framework import viewsets, permissions, filters, status
from django_filters.rest_framework import DjangoFilterBackend
from customers.models import Customer
from customers.serializers import CustomerSerializer
from notifications.services import create_notification
import io
from datetime import datetime
from openpyxl import load_workbook
from rest_framework.decorators import action
from rest_framework.response import Response


class CustomerViewSet(viewsets.ModelViewSet):
    """
    مدیریت کامل مشتریان (ایجاد، خواندن، بروزرسانی، حذف).

    این ViewSet تمام عملیات CRUD را برای مدل Customer فراهم می‌کند.
    کاربر احراز هویت‌شده باید باشد. هنگام ایجاد، فیلد created_by به‌طور
    خودکار با کاربر درخواست‌دهنده پر می‌شود.

    ویژگی‌ها:
        - فیلترپذیری بر اساس phone_number, full_name, car_model.
        - جستجوی متنی روی phone_number و full_name.
        - مرتب‌سازی پیش‌فرض بر اساس created_at نزولی.

    Endpoints:
        GET    /api/v1/customers/          - لیست مشتریان (با صفحه‌بندی)
        POST   /api/v1/customers/          - ایجاد مشتری جدید
        GET    /api/v1/customers/{id}/     - نمایش جزئیات یک مشتری
        PUT    /api/v1/customers/{id}/     - بروزرسانی کامل یک مشتری
        PATCH  /api/v1/customers/{id}/     - بروزرسانی جزئی یک مشتری
        DELETE /api/v1/customers/{id}/     - حذف یک مشتری

    Authentication:
        JWT (Bearer token)

    Permissions:
        IsAuthenticated
    """
    queryset = Customer.objects.all().order_by('-created_at')
    serializer_class = CustomerSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ['phone_number', 'full_name', 'car_model']
    search_fields = ['phone_number', 'full_name']
    ordering_fields = ['created_at', 'full_name']

    def perform_create(self, serializer):
        """
        هنگام ایجاد مشتری، کاربر ایجادکننده را به‌طور خودکار ثبت کن.

        Args:
            serializer: نمونهٔ CustomerSerializer با داده‌های معتبر.

        Returns:
            None
        """
        customer = serializer.save(created_by=self.request.user)
        create_notification(
            user=self.request.user,
            title="مشتری جدید",
            body=f"مشتری {customer.full_name} با شماره {customer.phone_number} با موفقیت ثبت شد."
        )
        serializer.save(created_by=self.request.user)


    @action(detail=False, methods=['post'], url_path='upload-excel')
    def upload_excel(self, request):
        """
        بارگذاری فایل اکسل و ایجاد گروهی مشتریان.

        ستون‌های مورد انتظار:
            - phone_number (اجباری)
            - full_name (اختیاری)
            - car_model (اختیاری)
            - car_usage_type (اختیاری)
            - birthday (اختیاری با فرمت YYYY-MM-DD)

        شماره‌های تکراری برای کاربر جاری نادیده گرفته می‌شوند.
        """
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'فایلی ارسال نشده است.'}, status=status.HTTP_400_BAD_REQUEST)

        if not file_obj.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'فقط فایل‌های Excel با پسوند xlsx یا xls پذیرفته می‌شوند.'},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(filename=io.BytesIO(file_obj.read()))
            ws = wb.active
        except Exception as e:
            return Response({'error': f'خطا در خواندن فایل: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        # تشخیص سرستون‌ها
        headers = [cell.value for cell in ws[1]]
        phone_idx = name_idx = car_model_idx = usage_idx = bday_idx = None
        for i, h in enumerate(headers):
            if h and h.lower().strip() == 'phone_number':
                phone_idx = i
            elif h and h.lower().strip() == 'full_name':
                name_idx = i
            elif h and h.lower().strip() == 'car_model':
                car_model_idx = i
            elif h and h.lower().strip() == 'car_usage_type':
                usage_idx = i
            elif h and h.lower().strip() == 'birthday':
                bday_idx = i

        if phone_idx is None:
            return Response({'error': 'ستون phone_number یافت نشد.'}, status=status.HTTP_400_BAD_REQUEST)

        created_count = 0
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) <= phone_idx or not row[phone_idx]:
                continue
            phone = str(row[phone_idx]).strip()
            full_name = str(row[name_idx]).strip() if name_idx is not None and row[name_idx] else ''
            car_model = str(row[car_model_idx]).strip() if car_model_idx is not None and row[car_model_idx] else ''
            car_usage = str(row[usage_idx]).strip() if usage_idx is not None and row[usage_idx] else ''
            birthday = None
            if bday_idx is not None and row[bday_idx]:
                try:
                    birthday = datetime.strptime(str(row[bday_idx]).strip(), '%Y-%m-%d').date()
                except ValueError:
                    errors.append(f'ردیف {row_idx}: فرمت تاریخ تولد نامعتبر.')
                    continue

            # جلوگیری از تکراری
            if Customer.objects.filter(phone_number=phone, created_by=request.user).exists():
                errors.append(f'ردیف {row_idx}: شماره {phone} قبلاً ثبت شده است.')
                continue

            Customer.objects.create(
                phone_number=phone,
                full_name=full_name,
                car_model=car_model,
                car_usage_type=car_usage,
                birthday=birthday,
                created_by=request.user
            )
            created_count += 1

        return Response({'created_count': created_count, 'errors': errors})
